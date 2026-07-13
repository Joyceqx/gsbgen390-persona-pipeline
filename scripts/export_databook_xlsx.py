"""Build the primary databook Excel:

Sheet 1 (Databook_1_Phase1A) — 5 models × 3 prompts × 12 items = 180 rows
                                 N=200 respondents (full Phase 1A cohort)
                                 n_eff target per cell = 400 (200 × 2 samples)
Sheet 2 (Databook_2_Anchor)  — 7 model choices × P0 × 12 items = 84 rows
                                 (5 cheap @ P0 + Anchor B + Anchor A)
                                 N=100 respondents (anchor cohort = first 100 of seed=42)
                                 n_eff target per cell = 200 (100 × 2 samples)

Cols (both): model | prompt | item | K | tx |
            em | mae_raw | w1 | n_eff |
            rnd_em | maj_em |
            lift_rnd | lift_maj |
            pred_distinct_codes | pred_top1_share

Row 2 of each data sheet is a legend row (short col explanations).
Highlighted columns (key metrics): em, mae_raw, lift_maj, pred_top1_share.
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/joyce/Developer/gsbgen390")
OUT_XLSX = ROOT / "outputs/raw_tables.xlsx"

sys.path.insert(0, str(ROOT / "src"))
from gss_pipeline import format_eval_question, load_taxonomy

tx = load_taxonomy()
PRIMARY = [it["id"] for it in tx["primary_eval"]["items"]]
META = {}
for it in tx["primary_eval"]["items"]:
    _, m = format_eval_question(it)
    META[it["id"]] = {"K": len(m["valid_codes"]), "codes": sorted(m["valid_codes"])}

PHASE1A = {
    "P0": ROOT / "outputs/gss_phase1_records_n200_qwen3-ma-deepseek-llama-4--kimi-k2-_seed42_P0.json",
    "P1": ROOT / "outputs/gss_phase1_records_n200_qwen3-ma-deepseek-llama-4--kimi-k2-_seed42_P1.json",
    "P2": ROOT / "outputs/gss_phase1_records_n200_qwen3-ma-deepseek-llama-4--kimi-k2-_seed42_P2.json",
}
ANCHOR_FILES = {
    "Anchor-B-R1ON":  ROOT / "outputs/anchor_r1on_n100.json",
    "Anchor-A-R1OFF": ROOT / "outputs/anchor_r1off_n100.json",
}
MODELS = ["qwen/qwen3-max", "deepseek/deepseek-v3.1-terminus",
          "meta-llama/llama-4-maverick", "moonshotai/kimi-k2-0905"]
MODELS_SHORT = {
    "qwen/qwen3-max": "q3m",
    "deepseek/deepseek-v3.1-terminus": "ds31",
    "meta-llama/llama-4-maverick": "l4m",
    "moonshotai/kimi-k2-0905": "kk2",
}
PROMPTS = ["P0", "P1", "P2"]

# Per-call cost (USD) — May 2026 OpenRouter / OpenAI snapshot, non-thinking models,
# ~950 input tokens + 1-2 output tokens per Phase 1A call.
# Source: src/select_phase1b_cell.py DEFAULT_COST_PER_CALL_USD + RESEARCH_DESIGN.md §11.
COST_PER_CALL = {
    "q3m":  9.8e-4,                # qwen/qwen3-max
    "ds31": 3.2e-4,                # deepseek/deepseek-v3.1-terminus
    "l4m":  1.8e-4,                # meta-llama/llama-4-maverick
    "kk2":  7.2e-4,                # moonshotai/kimi-k2-0905
    "rnd":  0.0,                    # analytical, no API call
    "Anchor-B-R1ON":  2.4e-3,       # openai/gpt-4o-2024-08-06 (est)
    "Anchor-A-R1OFF": 2.4e-3,       # openai/gpt-4o-2024-08-06 (est)
}

BOOTSTRAP_B = 2000   # bootstrap replicates for 95% CI (B=10000 in formal §7; B=2000 here for speed, ±1% noise)
BOOTSTRAP_SEED = 42

# -------- Load cohort + data --------
def first_n(p, n):
    with open(p) as f: r = json.load(f)
    seen = []
    for x in r:
        if x["respondent_id"] not in seen: seen.append(x["respondent_id"])
        if len(seen) >= n: break
    return set(seen[:n])

# Two cohorts:
#   FULL_200  = all 200 Phase 1A respondents (for Databook 1)
#   FIRST_100 = first 100 respondents (anchor-paired, for Databook 2)
COHORT_FULL_200 = first_n(PHASE1A["P0"], 200)
COHORT_FIRST_100 = first_n(PHASE1A["P0"], 100)
assert len(COHORT_FULL_200) == 200, f"expected 200, got {len(COHORT_FULL_200)}"
assert COHORT_FIRST_100.issubset(COHORT_FULL_200), "first 100 must be subset of full 200"

def load_(p, cohort=None):
    with open(p) as f: r = json.load(f)
    return [x for x in r if cohort is None or x["respondent_id"] in cohort]

# Databook 1 uses full 200; Databook 2 cheap rows use first 100 (paired with anchor)
phase1a_full = {p: load_(path, COHORT_FULL_200) for p, path in PHASE1A.items()}
phase1a_100  = {p: load_(path, COHORT_FIRST_100) for p, path in PHASE1A.items()}
anchor = {l: [r for r in load_(p) if r.get("condition") == "full"] for l, p in ANCHOR_FILES.items()}

# -------- Aggregate metrics for one (records, item) slice --------
def aggregate(records, item_id):
    """Returns aggregate metrics + per-respondent stats for bootstrap.

    mae_normalized follows §7 selector convention: per sample, distance is
      - parse_fail   → 1.0  (maximum, max distance)
      - ordinal ok   → abs_err / (max_code - min_code)
      - categorical  → (1 - cat_match) / 1   [for K=2; for K>2 cat, same 0/1]
    Then averaged across ALL samples (including parse_fail), so the metric is
    conservative — a high parse_fail_rate inflates mae_normalized.
    """
    codes = META[item_id]["codes"]
    code_range = max(codes) - min(codes) if len(codes) > 1 else 1
    n_total = n_ok = n_exact = n_w1 = n_parse_fail = 0
    mae_raw_sum = 0.0; mae_raw_count = 0
    mae_norm_sum = 0.0  # over all n_total samples
    is_ord = False
    preds = []
    # Per-respondent for bootstrap
    by_rid_em = defaultdict(list)
    by_rid_mae_norm = defaultdict(list)

    for r in records:
        rid = r["respondent_id"]
        for s in r["per_item_scores"].get(item_id, []):
            if s.get("skipped_missing_truth"):
                continue   # no truth = exclude from denominator
            n_total += 1
            if s.get("parse_fail"):
                n_parse_fail += 1
                mae_norm_sum += 1.0           # §7: parse fail → max distance
                by_rid_em[rid].append(0)
                by_rid_mae_norm[rid].append(1.0)
                continue
            n_ok += 1
            ae = s.get("abs_err"); cm = s.get("cat_match"); w1 = s.get("within1")
            preds.append(s.get("persona_code"))
            if ae is not None:
                is_ord = True
                if ae == 0:
                    n_exact += 1
                    by_rid_em[rid].append(1)
                    by_rid_mae_norm[rid].append(0.0)
                else:
                    by_rid_em[rid].append(0)
                    by_rid_mae_norm[rid].append(ae / code_range)
                if w1: n_w1 += 1
                mae_raw_sum += ae; mae_raw_count += 1
                mae_norm_sum += ae / code_range
            elif cm is not None:
                if cm == 1:
                    n_exact += 1
                    by_rid_em[rid].append(1)
                    by_rid_mae_norm[rid].append(0.0)
                else:
                    by_rid_em[rid].append(0)
                    by_rid_mae_norm[rid].append(1.0)
                    mae_norm_sum += 1.0
    em = n_exact / n_ok if n_ok else None
    mae_raw = mae_raw_sum / mae_raw_count if mae_raw_count else None
    w1 = n_w1 / n_ok if n_ok and is_ord else None
    mae_normalized = mae_norm_sum / n_total if n_total else None
    parse_fail_rate = n_parse_fail / n_total if n_total else None
    # output diversity diagnostic
    pred_dist = Counter(p for p in preds if p is not None)
    K = META[item_id]["K"]
    distinct = len(pred_dist)
    top1_share = max(pred_dist.values()) / sum(pred_dist.values()) if pred_dist else None
    # per-respondent values (mean over samples) for bootstrap
    per_resp_em = [mean(v) for v in by_rid_em.values() if v]
    per_resp_mae_norm = [mean(v) for v in by_rid_mae_norm.values() if v]
    # truth values (one per respondent)
    truths = []
    for r in records:
        for s in r["per_item_scores"].get(item_id, []):
            if s.get("skipped_missing_truth"): continue
            t = s.get("truth")
            if t is not None:
                truths.append(t); break
    return {
        "em": em, "mae_raw": mae_raw, "w1": w1,
        "n_eff": n_ok, "n_total": n_total,
        "mae_normalized": mae_normalized,
        "parse_fail_rate": parse_fail_rate,
        "is_ord": is_ord,
        "pred_distinct": distinct, "pred_K": K,
        "pred_top1_share": top1_share,
        "per_resp_em": per_resp_em,
        "per_resp_mae_norm": per_resp_mae_norm,
        "truths": truths,
    }


def bootstrap_ci(per_respondent_values, B=BOOTSTRAP_B, alpha=0.05, seed=BOOTSTRAP_SEED):
    """Respondent-level resample bootstrap CI on the mean."""
    if not per_respondent_values: return None, None
    arr = np.asarray(per_respondent_values, dtype=float)
    n = len(arr)
    if n < 2: return None, None
    rng = np.random.default_rng(seed)
    idx = rng.integers(0, n, size=(B, n))
    boots = arr[idx].mean(axis=1)
    return float(np.percentile(boots, 100 * alpha / 2)), float(np.percentile(boots, 100 * (1 - alpha / 2)))

# Random baseline (per item): 1/K
def rnd_em(item_id):
    return 1.0 / META[item_id]["K"]

# Majority-class baseline (per item, from cohort truth distribution)
def maj_em(truths):
    if not truths: return None
    counts = Counter(t for t in truths if t is not None)
    if not counts: return None
    return max(counts.values()) / sum(counts.values())

# Detect treatment per item from data
treatments = {}
for item in PRIMARY:
    is_ord = False
    for r in phase1a_full["P0"]:
        for s in r["per_item_scores"].get(item, []):
            if s.get("abs_err") is not None: is_ord = True; break
            elif s.get("cat_match") is not None: is_ord = False; break
        if r["per_item_scores"].get(item): break
    treatments[item] = "ord" if is_ord else "cat"

# -------- Build Databook 1: Phase 1A (5 models × 3 prompts × 12 items) --------
DB1_HEADERS = [
    "model", "prompt", "item", "K", "tx",
    "em", "em_ci_lo", "em_ci_hi",
    "mae_raw", "mae_normalized", "mae_norm_ci_lo", "mae_norm_ci_hi",
    "w1",
    "n_eff", "parse_fail_rate",
    "rnd_em", "maj_em",
    "lift_rnd", "lift_maj",
    "pred_distinct_codes", "pred_top1_share",
    "cost_per_call",
]

def lift(em, base):
    if em is None or base is None: return None
    if base >= 1: return None
    return (em - base) / (1 - base)

db1_rows = []
# Majority baseline for Databook 1 = computed from full 200-respondent truths
maj_em_db1 = {}
for item in PRIMARY:
    any_recs = [r for r in phase1a_full["P0"] if r["model"] == MODELS[0]]
    s = aggregate(any_recs, item)
    maj_em_db1[item] = maj_em(s["truths"])

# Helper: cache rnd row's n_eff by inheriting from q3m·P0·item (same ballot rotation)
rnd_n_eff_db1 = {}
for item in PRIMARY:
    recs = [r for r in phase1a_full["P0"] if r["model"] == MODELS[0]]
    s_ref = aggregate(recs, item)
    rnd_n_eff_db1[item] = s_ref["n_eff"]

# Add 4 cheap models × 3 prompts × 12 items (cohort = full 200)
for model in MODELS:
    short = MODELS_SHORT[model]
    for prompt in PROMPTS:
        recs = [r for r in phase1a_full[prompt] if r["model"] == model]
        for item in PRIMARY:
            s = aggregate(recs, item)
            em_lo, em_hi = bootstrap_ci(s["per_resp_em"])
            mn_lo, mn_hi = bootstrap_ci(s["per_resp_mae_norm"])
            rnd = rnd_em(item)
            maj = maj_em_db1[item]
            db1_rows.append({
                "model": short, "prompt": prompt, "item": item,
                "K": META[item]["K"], "tx": treatments[item],
                "em": s["em"], "em_ci_lo": em_lo, "em_ci_hi": em_hi,
                "mae_raw": s["mae_raw"],
                "mae_normalized": s["mae_normalized"],
                "mae_norm_ci_lo": mn_lo, "mae_norm_ci_hi": mn_hi,
                "w1": s["w1"],
                "n_eff": s["n_eff"], "parse_fail_rate": s["parse_fail_rate"],
                "rnd_em": rnd, "maj_em": maj,
                "lift_rnd": lift(s["em"], rnd), "lift_maj": lift(s["em"], maj),
                "pred_distinct_codes": s["pred_distinct"],
                "pred_top1_share": s["pred_top1_share"],
                "cost_per_call": COST_PER_CALL[short],
            })

# Add 5th model = random baseline × 3 prompts × 12 items (analytical)
for prompt in PROMPTS:
    for item in PRIMARY:
        rnd = rnd_em(item)
        maj = maj_em_db1[item]
        K = META[item]["K"]
        db1_rows.append({
            "model": "rnd", "prompt": prompt, "item": item,
            "K": K, "tx": treatments[item],
            "em": rnd, "em_ci_lo": rnd, "em_ci_hi": rnd,
            "mae_raw": None,
            "mae_normalized": 1 - rnd,  # uniform random's expected normalized distance
            "mae_norm_ci_lo": 1 - rnd, "mae_norm_ci_hi": 1 - rnd,
            "w1": None,
            "n_eff": rnd_n_eff_db1[item],  # inherit ballot context from q3m·P0
            "parse_fail_rate": 0.0,
            "rnd_em": rnd, "maj_em": maj,
            "lift_rnd": 0.0, "lift_maj": lift(rnd, maj),
            "pred_distinct_codes": K,
            "pred_top1_share": 1.0 / K,
            "cost_per_call": COST_PER_CALL["rnd"],
        })

print(f"Databook 1: {len(db1_rows)} rows (expected 5×3×12 = 180)")

# -------- Build Databook 2: Anchor comparison at P0 (7 models × 12 items) --------
DB2_HEADERS = list(DB1_HEADERS)  # identical schema

db2_rows = []
# Majority baseline for Databook 2 = computed from first-100 truths (matches anchor cohort)
maj_em_db2 = {}
rnd_n_eff_db2 = {}
for item in PRIMARY:
    any_recs = [r for r in phase1a_100["P0"] if r["model"] == MODELS[0]]
    s = aggregate(any_recs, item)
    maj_em_db2[item] = maj_em(s["truths"])
    rnd_n_eff_db2[item] = s["n_eff"]

# 5 cheap models @ P0 (4 actual + random) — cohort = first 100
for model in MODELS:
    short = MODELS_SHORT[model]
    recs = [r for r in phase1a_100["P0"] if r["model"] == model]
    for item in PRIMARY:
        s = aggregate(recs, item)
        em_lo, em_hi = bootstrap_ci(s["per_resp_em"])
        mn_lo, mn_hi = bootstrap_ci(s["per_resp_mae_norm"])
        rnd = rnd_em(item); maj = maj_em_db2[item]
        db2_rows.append({
            "model": short, "prompt": "P0", "item": item,
            "K": META[item]["K"], "tx": treatments[item],
            "em": s["em"], "em_ci_lo": em_lo, "em_ci_hi": em_hi,
            "mae_raw": s["mae_raw"],
            "mae_normalized": s["mae_normalized"],
            "mae_norm_ci_lo": mn_lo, "mae_norm_ci_hi": mn_hi,
            "w1": s["w1"],
            "n_eff": s["n_eff"], "parse_fail_rate": s["parse_fail_rate"],
            "rnd_em": rnd, "maj_em": maj,
            "lift_rnd": lift(s["em"], rnd), "lift_maj": lift(s["em"], maj),
            "pred_distinct_codes": s["pred_distinct"],
            "pred_top1_share": s["pred_top1_share"],
            "cost_per_call": COST_PER_CALL[short],
        })

# Random model (P0) — analytical
for item in PRIMARY:
    rnd = rnd_em(item); maj = maj_em_db2[item]
    K = META[item]["K"]
    db2_rows.append({
        "model": "rnd", "prompt": "P0", "item": item,
        "K": K, "tx": treatments[item],
        "em": rnd, "em_ci_lo": rnd, "em_ci_hi": rnd,
        "mae_raw": None,
        "mae_normalized": 1 - rnd,
        "mae_norm_ci_lo": 1 - rnd, "mae_norm_ci_hi": 1 - rnd,
        "w1": None,
        "n_eff": rnd_n_eff_db2[item], "parse_fail_rate": 0.0,
        "rnd_em": rnd, "maj_em": maj,
        "lift_rnd": 0.0, "lift_maj": lift(rnd, maj),
        "pred_distinct_codes": K,
        "pred_top1_share": 1.0 / K,
        "cost_per_call": COST_PER_CALL["rnd"],
    })

# 2 anchor conditions @ P0
for label, recs in anchor.items():
    for item in PRIMARY:
        s = aggregate(recs, item)
        em_lo, em_hi = bootstrap_ci(s["per_resp_em"])
        mn_lo, mn_hi = bootstrap_ci(s["per_resp_mae_norm"])
        rnd = rnd_em(item); maj = maj_em_db2[item]
        db2_rows.append({
            "model": label, "prompt": "P0", "item": item,
            "K": META[item]["K"], "tx": treatments[item],
            "em": s["em"], "em_ci_lo": em_lo, "em_ci_hi": em_hi,
            "mae_raw": s["mae_raw"],
            "mae_normalized": s["mae_normalized"],
            "mae_norm_ci_lo": mn_lo, "mae_norm_ci_hi": mn_hi,
            "w1": s["w1"],
            "n_eff": s["n_eff"], "parse_fail_rate": s["parse_fail_rate"],
            "rnd_em": rnd, "maj_em": maj,
            "lift_rnd": lift(s["em"], rnd), "lift_maj": lift(s["em"], maj),
            "pred_distinct_codes": s["pred_distinct"],
            "pred_top1_share": s["pred_top1_share"],
            "cost_per_call": COST_PER_CALL[label],
        })

print(f"Databook 2: {len(db2_rows)} rows (expected 7×12 = 84)")

# -------- Build Excel --------
wb = Workbook()
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
LEGEND_FILL = PatternFill("solid", fgColor="EEEEEE")
KEY_METRIC_HEADER_FILL = PatternFill("solid", fgColor="C6EFCE")  # green
KEY_METRIC_DATA_FILL = PatternFill("solid", fgColor="EBF7E6")    # pale green
HEADER_FONT = Font(bold=True)
LEGEND_FONT = Font(italic=True, color="555555", size=9)
ROW_FILL = PatternFill("solid", fgColor="F5F5F5")

# Short legend strings for each col (row 2 of data sheets)
LEGEND = {
    "model": "cheap LLM / rnd / Anchor",
    "prompt": "P0 / P1 / P2",
    "item": "GSS variable",
    "K": "# valid codes",
    "tx": "ord / cat",
    "em": "ExactM (Park metric)",
    "em_ci_lo": "EM 95% CI low (bootstrap B=2k)",
    "em_ci_hi": "EM 95% CI high",
    "mae_raw": "Mean abs err (ord only)",
    "mae_normalized": "§7 selector metric (parse_fail = 1.0) ★",
    "mae_norm_ci_lo": "MAE_norm 95% CI low",
    "mae_norm_ci_hi": "MAE_norm 95% CI high",
    "w1": "Within ±1 (ord only)",
    "n_eff": "Effective samples (parsed)",
    "parse_fail_rate": "Parse fail fraction ★",
    "rnd_em": "Uniform baseline = 1/K",
    "maj_em": "Majority-class baseline",
    "lift_rnd": "(em−rnd) / (1−rnd)",
    "lift_maj": "(em−maj) / (1−maj) ★",
    "pred_distinct_codes": "# distinct codes used",
    "pred_top1_share": "Top-1 freq (>.95 = collapse)",
    "cost_per_call": "USD per API call ★",
}

# Columns to highlight as KEY METRICS to pay attention to (green fill)
KEY_METRIC_COLS = {"em", "mae_normalized", "lift_maj", "parse_fail_rate", "pred_top1_share", "cost_per_call"}

def style_data_sheet(ws, headers):
    """Apply header row, legend row, freeze, sizing, stripes, highlights."""
    # Identify key metric column indices
    key_idx = {i + 1 for i, h in enumerate(headers) if h in KEY_METRIC_COLS}
    # Row 1: headers
    for ci, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if ci in key_idx:
            cell.fill = KEY_METRIC_HEADER_FILL
        else:
            cell.fill = HEADER_FILL
    # Row 2: legend
    for ci, cell in enumerate(ws[2], 1):
        cell.font = LEGEND_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = LEGEND_FILL
        cell.border = BORDER
    ws.row_dimensions[2].height = 28
    # Data rows: stripe + highlight key cols
    for r in range(3, ws.max_row + 1):
        is_stripe = (r - 3) % 2 == 1
        for ci, cell in enumerate(ws[r], 1):
            if ci in key_idx:
                cell.fill = KEY_METRIC_DATA_FILL
            elif is_stripe:
                cell.fill = ROW_FILL

def autosize(ws, max_w=22):
    for ci in range(1, ws.max_column + 1):
        m = 8
        for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
            v = row[0]
            if v is not None: m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w, m + 2)

def round_v(v, p=4):
    if v is None: return None
    if isinstance(v, int): return v
    return round(v, p)

def emit_row(ws, row_dict, headers):
    vals = []
    for h in headers:
        v = row_dict.get(h)
        if v is None:
            vals.append(None)
        elif isinstance(v, int):
            vals.append(v)
        elif isinstance(v, float):
            # cost_per_call needs more precision
            vals.append(round(v, 6) if h == "cost_per_call" else round(v, 4))
        else:
            vals.append(v)
    ws.append(vals)

# Sheet 1: Databook 1
ws = wb.active
ws.title = "Databook_1_Phase1A"
ws.append(DB1_HEADERS)
ws.append([LEGEND[h] for h in DB1_HEADERS])
for row in db1_rows:
    emit_row(ws, row, DB1_HEADERS)
ws.freeze_panes = "F3"
autosize(ws)
style_data_sheet(ws, DB1_HEADERS)

# Sheet 2: Databook 2
ws = wb.create_sheet("Databook_2_Anchor")
ws.append(DB2_HEADERS)
ws.append([LEGEND[h] for h in DB2_HEADERS])
for row in db2_rows:
    emit_row(ws, row, DB2_HEADERS)
ws.freeze_panes = "F3"
autosize(ws)
style_data_sheet(ws, DB2_HEADERS)

# Sheet 3: README
ws = wb.create_sheet("README")
README = [
    ["Phase 1A + Anchor — Primary Databook"],
    [""],
    ["Generated: 2026-06-01"],
    [""],
    ["▶ Background"],
    ["", "GSBGEN390 thesis project (Stanford GSB, advisor: Prof. Mohsen Bayati). Methodological study benchmarking against Park et al. 2024 v2 (arXiv:2411.10109)."],
    ["", ""],
    ["", "Research question — can CHEAP LLMs (per-call cost ~$0.0002 – $0.001) substitute for FRONTIER GPT-4o (~$0.0024/call) when simulating GSS 2024 attitude responses from demographic + behavioral persona inputs?"],
    ["", ""],
    ["", "Setup — 12 stratified GSS attitude items (politics, abortion, gender, race, institutional trust, redistribution, personal finance). Persona = each respondent's GSS demographic + behavioral + psychological features, MINUS same-battery siblings (R1 leakage exclusion)."],
    ["", ""],
    ["", "Phase 1A answers two questions:"],
    ["", "  Q1  How much accuracy is lost moving from GPT-4o to cheap models?  → Databook 2 (cheap vs anchor at P0, paired N=100)"],
    ["", "  Q2  Which (model, prompt) cell wins for Phase 1B scaling?  → Databook 1 (full 5×3×12 factorial, N=200)"],
    ["", ""],
    ["", "Anchor (Databook 2) also reports Park v2 Strategy A direct comparison (Anchor-A-R1OFF row): GPT-4o without R1 protocol, matches Park's surveys-only protocol exactly."],
    [""],
    ["▶ Cost — why \"cheap\""],
    ["", "Per-call USD cost (May 2026 OpenRouter / OpenAI direct snapshot; non-thinking models; ~950 input + 1-2 output tokens per Phase 1A call):"],
    ["", ""],
    ["", "  Model                                $/call       vs GPT-4o"],
    ["", "  ─────────────────────────────────────────────────────────────────"],
    ["", "  openai/gpt-4o-2024-08-06             $0.00240      1.0× (reference)"],
    ["", "  qwen/qwen3-max                       $0.00098      2.4× cheaper"],
    ["", "  moonshotai/kimi-k2-0905              $0.00072      3.3× cheaper"],
    ["", "  deepseek/deepseek-v3.1-terminus      $0.00032      7.5× cheaper"],
    ["", "  meta-llama/llama-4-maverick          $0.00018      13× cheaper"],
    ["", ""],
    ["", "Phase 1A total cost: ~$14 (4 cheap models × 3 prompts × N=200 × 12 items × n=2). Anchor cost: ~$48 (GPT-4o × P0 × N=100 × 12 items × 2 R1 conditions × n=2)."],
    ["", ""],
    ["", "If a cheap (model, prompt) cell matches GPT-4o accuracy → 2-13× cost reduction translates to Phase 1B scaling at N=3,309 (full GSS 2024 cross-section, the population-level inferential goal)."],
    [""],
    ["▶ Sheet structure"],
    ["  Databook_1_Phase1A", "5 model × 3 prompt × 12 item = 180 rows. Cohort = N=200 respondents (full Phase 1A draw). n_eff target per cell = 400 (200 × 2 samples)."],
    ["  Databook_2_Anchor",  "7 model × P0 × 12 item = 84 rows. Cohort = N=100 (anchor cohort = first 100 of seed=42). n_eff target per cell = 200 (100 × 2 samples)."],
    ["  README",             "This sheet."],
    [""],
    ["★ KEY METRICS to focus on (highlighted GREEN in data sheets)"],
    ["  em",                "ExactM — Park v2's headline metric. Use for direct external comparison."],
    ["  mae_normalized",    "RESEARCH_DESIGN.md §7 selector's primary metric. Per-sample distance / item range, with parse_fail counted as 1.0 (max distance). Lower is better."],
    ["  lift_maj",          "Lift over majority-class baseline. Honest 'is the model learning?' test. > 0 = model beats 'always predict the most common answer'. Negative = model worse than 'just predict mode'."],
    ["  parse_fail_rate",   "Fraction of samples the model failed to return a parseable code. §7 DQ-1: cells with parse_fail_rate > 0.10 are disqualified."],
    ["  pred_top1_share",   "Top-1 prediction frequency. > 0.95 with K≥3 = MODEL COLLAPSE (outputs same code regardless of persona). For K=2, look at pred_distinct_codes=1 instead."],
    ["  cost_per_call",     "USD per API call. Drives §7 cost tiebreak among CI-overlapping cells. l4m cheapest at $0.00018/call."],
    [""],
    ["Other column meanings"],
    ["  model",                 "Cheap: q3m / ds31 / l4m / kk2; rnd = uniform random baseline; Anchor-B-R1ON / Anchor-A-R1OFF (only in Databook 2)"],
    ["  prompt",                "P0 / P1 / P2 in Databook 1; all P0 in Databook 2"],
    ["  item",                  "GSS variable name (12 primary_eval items)"],
    ["  K",                     "# valid response codes for this item (e.g., POLVIEWS K=7, binary items K=2)"],
    ["  tx",                    "ord = ordinal Likert; cat = categorical (mostly binary)"],
    ["  em_ci_lo / em_ci_hi",   "95% bootstrap CI on em. Respondent-level resampling, B=2,000. CIs overlapping between two cells = those cells are not statistically distinguishable."],
    ["  mae_raw",                "Raw mean abs err in original code units. Ordinal items only (empty for cat). Excludes parse_fail. Use mae_normalized for §7 selector."],
    ["  mae_norm_ci_lo / hi",    "95% bootstrap CI on mae_normalized (B=2,000)."],
    ["  w1 (Within±1)",         "Fraction within 1 Likert point; ordinal only. Soft accuracy."],
    ["  n_eff",                 "Effective samples (parsed). DB1 target = 400; DB2 target = 200. Lower than target = GSS ballot rotation dropped respondents for this item. Random rows inherit n_eff from q3m·P0·item for context."],
    ["  rnd_em",                "Uniform random baseline = 1/K. Over-optimistic — real human truths are skewed, so 'always predict majority' beats uniform."],
    ["  maj_em",                "Majority-class baseline = max share in the cohort's truth distribution for this item. Honest baseline."],
    ["  lift_rnd",              "(em − rnd_em) / (1 − rnd_em). 0 = no better than uniform random; 1 = perfect."],
    ["  pred_distinct_codes",   "How many distinct codes the model used (out of K). If < K, model is partially collapsing. = 1 means total collapse."],
    [""],
    ["Cheap model short codes (with $/call for context)"],
    ["  q3m",  "qwen/qwen3-max (Alibaba) — $0.00098/call, 2.4× cheaper than GPT-4o"],
    ["  ds31", "deepseek/deepseek-v3.1-terminus (DeepInfra) — $0.00032/call, 7.5× cheaper"],
    ["  l4m",  "meta-llama/llama-4-maverick (DeepInfra) — $0.00018/call, 13× cheaper"],
    ["  kk2",  "moonshotai/kimi-k2-0905 (Novita) — $0.00072/call, 3.3× cheaper"],
    ["  rnd",  "uniform random over K valid codes (analytical, no API call, $0)"],
    [""],
    ["Anchor labels (Databook 2 only) — $0.00240/call each"],
    ["  Anchor-B-R1ON",  "GPT-4o + R1 battery exclusion (same protocol as Phase 1A cheap rows). The internal-frontier comparison."],
    ["  Anchor-A-R1OFF", "GPT-4o + no R1 exclusion (Park Strategy A direct). The external cross-study reference."],
    [""],
    ["Caveats (also see notes/park_comparability.md)"],
    ["  • FEPOL / RACDIF1 have smaller n_eff (≈ 50 / ≈ 60 in DB2) due to GSS ballot rotation. Paired tests on these items have wider CIs."],
    ["  • llama on CONFINAN: pred_distinct_codes = 1/3 with top1_share = 100% → MODEL COLLAPSE. EM is luck not learning."],
    ["  • deepseek on CONFINAN: pred_top1_share ≈ 97% → near-collapse, same caveat."],
    ["  • rnd_em (uniform) is over-optimistic because human truths are skewed; use maj_em / lift_maj for honest baseline."],
    ["  • For K=2 binary items, pred_top1_share > 0.95 is not automatic collapse (binary skew is common); flag pred_distinct_codes = 1 instead."],
    [""],
    ["§7 selector inputs status (per RESEARCH_DESIGN.md §7)"],
    ["  ✓ mae_normalized",     "Now in databook (parse_fail=1.0 convention)"],
    ["  ✓ parse_fail_rate",    "Now in databook (DQ-1 threshold ≤ 10%)"],
    ["  ✓ bootstrap CIs",      "Now in databook (B=2,000; formal §7 uses B=10,000)"],
    ["  ✓ cost_per_call",      "Now in databook"],
    ["  ✗ dq3_variance_ratio", "NOT in databook. DQ-3 (per-item model variance ÷ human variance < 0.30) requires `primary_eval_human_variance_2024.json` cross-reference. Compute externally if needed for §7 review."],
    ["  Note",                 "Selector decision itself (which cell wins Phase 1B) is intentionally NOT in this databook — discuss with advisor first."],
]
for r in README:
    ws.append(r)
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 100
ws["A1"].font = Font(bold=True, size=14)

wb.save(OUT_XLSX)
print(f"\n✓ wrote {OUT_XLSX}")
print(f"  3 sheets: Databook_1_Phase1A · Databook_2_Anchor · README")
